# # 脚本写于2022年9月，由于iGEM可能会使用新的网址或者修改元素的Xpath
# # 因此如果程序无法正确运行，请去检查网页链接和元素的Xpath是否改变
# # author: Xiaoping Yu

# # 测试，请把亮着的代码注释掉，将这段代码的注释取消
# # 尝试单独运行这段代码，如果运行不了则是环境安装的问题
# # from selenium import webdriver
# # driver = webdriver.Chrome()
# # driver.get('http://www.baidu.com')

#  以下为上传composite parts的代码，可以加一个错误验证(如过表单填写失败则中止程序)
# （也可以在parts列表查看是否有缺录，缺录一般是因为表格数据不正确，未成功提交）
#
# 确保已安装必要的第三方库
# 如果你使用的是pycharm，可以在File=>Setting=>PythonInterpreter=>pip中搜索selenium和openpyxl两个库进行install
from selenium import webdriver
from selenium.webdriver.support.ui import Select
import time
import openpyxl
from selenium.webdriver.common.by import By

js="window.open('{}','_blank');"

# 打开D盘的data数据表，也可以将excel表放在其他路径
wb = openpyxl.load_workbook(r'd:\parts.xlsx')
ws = wb.active
r = ws.max_row

# 指定浏览器为chrome，需先把selenium的chromeDriver放在python安装目录和chrome的安装目录
# 注意下载的chromeDriver版本需要与当前chrome浏览器版本对应
driver = webdriver.Chrome()
driver.get(r'https://old.igem.org/Login2')
time.sleep(1)

# 输入账户信息
# 代替为自己的用户名和密码
time.sleep(0.5)
driver.find_element(By.XPATH, '//*[@id="name_and_pass"]/input[1]').send_keys('icebear')
driver.find_element(By.XPATH, '//*[@id="name_and_pass"]/input[2]').send_keys('ccd6a890')
driver.find_element(By.XPATH, '//*[@id="login_form"]/input[2]').click()
time.sleep(0.5)

driver.execute_script(js.format('http://parts.igem.org/cgi/partsdb/add_part_c.cgi'))
driver.switch_to.window(driver.window_handles[-1])    #切换到最新页面
time.sleep(0.5)
driver.maximize_window()
time.sleep(1)

for i in range(2, r + 1):
    # 将excel各对应单元格的值输入页面
    # Attention!!
    # excel格式：name(第一列)、描述（第四列）、sub（第三列），可以自己改,不要写错
    time.sleep(1)
    # allow edit
    driver.find_element(By.XPATH, '//*[@id="table_info"]/tbody/tr[4]/td[1]/input').click()
    # 输入part name
    driver.find_element(By.NAME, 'part_name').send_keys(ws.cell(i,1).value)
    # 选择类型
    Select(driver.find_element(By.ID, "type")).select_by_value('Composite')
    # 输入描述
    driver.find_element(By.NAME, 'short_description').send_keys(ws.cell(i,4).value)
    # 输入三个none
    driver.find_element(By.NAME, 'long_description').send_keys('none')
    driver.find_element(By.NAME, 'source').send_keys('none')
    driver.find_element(By.NAME, 'notes').send_keys('none')
    # 输入sub
    driver.find_element(By.NAME, 'subparts').send_keys(ws.cell(i,3).value)
    # 结束,休息并返回
    driver.find_element(By.NAME, 'proceed').click()
    time.sleep(1)
    driver.close()
    time.sleep(1)
    driver.switch_to.window(driver.window_handles[-1])  # 切换到最新页面
    driver.execute_script(js.format('http://parts.igem.org/cgi/partsdb/add_part_c.cgi'))
    driver.switch_to.window(driver.window_handles[-1])  # 切换到最新页面

# 批量edit所有sub部分的代码
# 也可以修改成其他内容的edit代码
# 如果要使用这段代码，请把下面的注释去掉，同时将上面亮着的代码注释掉

# 确保已安装必要的第三方库
# 如果你使用的是pycharm，可以在File=>Setting=>PythonInterpreter=>pip中搜索selenium和openpyxl两个库进行install
#
# from selenium import webdriver
# import time
# import openpyxl
# from selenium.webdriver.common.by import By
#
#
# js="window.open('{}','_blank');"
#
# # 打开D盘的data数据表，也可以将excel表放在其他的盘
# wb = openpyxl.load_workbook(r'd:\parts.xlsx')
# ws = wb.active
# r = ws.max_row
#
# # 指定浏览器为chrome，需先把selenium的chromeDriver放在python安装目录
# driver = webdriver.Chrome()
# driver.get(r'https://old.igem.org/Login2')
# time.sleep(1)
#
# # 输入账户信息
# # 代替为自己的用户名和密码
# driver.find_element(By.XPATH, '//*[@id="name_and_pass"]/input[1]').send_keys('icebear')
# driver.find_element(By.XPATH, '//*[@id="name_and_pass"]/input[2]').send_keys('ccd6a890')
# driver.find_element(By.XPATH, '//*[@id="login_form"]/input[2]').click()
# time.sleep(0.5)
#
# driver.switch_to.window(driver.window_handles[-1])    #切换到最新页面
# driver.maximize_window()
#
# for i in range(2, r + 1):
#
#     temp='http://parts.igem.org/cgi/partsdb/edit_seq.cgi?part='+str(ws.cell(i,1).value)
#     driver.execute_script(js.format(temp))
#     driver.switch_to.window(driver.window_handles[-1])  # 切换到最新页面
#     time.sleep(1)
#     driver.find_element(By.XPATH, '//*[@id="form_dna"]/table[1]/tbody/tr/td[3]/a').click()
#     # 清空sub
#     driver.find_element(By.XPATH, '//*[@id="table_dna"]/tbody/tr[2]/td/textarea').clear()
#     # 输入sub
#     driver.find_element(By.XPATH, '//*[@id="table_dna"]/tbody/tr[2]/td/textarea').send_keys(ws.cell(i,3).value)
#     # 结束,休息并返回
#     driver.find_element(By.XPATH, '//*[@id="form_dna"]/table[1]/tbody/tr/td[4]/span[1]').click()
#     time.sleep(0.5)
#     # 对错误进行检查
#     target = driver.find_element(By.XPATH, '//*[@id="table_dna"]/tbody/tr[1]/td')
#     gbaColor = target.value_of_css_property("background-color")
#     #说明出现错误
#     if(str(gbaColor)=='rgba(255, 136, 136, 1)'):
#         print('wrong')
#         break
#     time.sleep(0.5)
#     driver.close()
#     driver.switch_to.window(driver.window_handles[-1])  # 切换到最新页面
